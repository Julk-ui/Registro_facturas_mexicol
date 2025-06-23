import pandas as pd
from openpyxl import load_workbook
import os

def estandarizar_columnas(df):
    df.columns = (
        df.columns
        .astype(str)
        .str.upper()
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    return df



def extract_data_from_excel(file_path):
    print(f"[INFO] Leyendo Excel: {file_path}")

    # Cargar encabezados desde hoja "cta cobro"
    df_raw = pd.read_excel(file_path, sheet_name="cta cobro", header=None)
    header_row = None
    for i, row in df_raw.iterrows():
        if row.astype(str).str.contains("PRODUCTO", case=False).any():
            header_row = i
            break

    if header_row is None:
        raise ValueError("No se encontró la fila de encabezado de productos")

    print(f"[INFO] Encabezado de productos encontrado en fila {header_row + 1}")
    df = pd.read_excel(file_path, sheet_name="cta cobro", header=header_row)
    df = df.rename(columns=lambda x: str(x).strip())

    # Renombrar columnas
    rename_map = {
        "CANTIDAD": "CANT. KLS",
        "DESCRIPCIÓN": "UNIDAD",
        "VR. UNITARIO": "VALOR X KL",
        "VR. UNITARIO ": "VALOR X KL",
        "VALOR X KI": "VALOR X KL",
        "TOTAL": "VALOR TOTAL",
    }
    df = df.rename(columns=rename_map)

    if "PRODUCTO" not in df.columns:
        raise ValueError("La columna 'PRODUCTO' no fue detectada")
    df = df[df["PRODUCTO"].notna()].copy()

    # Abrimos con openpyxl para celdas clave
    wb = load_workbook(file_path, data_only=True)
    ws = wb["cta cobro"]

    # FACTURA desde H6, FECHA desde H7
    cuenta_cobro = str(ws["H6"].value).strip() if ws["H6"].value else ""
    fecha = str(ws["H7"].value).strip() if ws["H7"].value else ""
    df["NUMERO FACTURA"] = cuenta_cobro
    df["FECHA"] = fecha
    print(f"[INFO] FACTURA: {cuenta_cobro} | FECHA: {fecha}")

    # FORMA DE PAGO desde B33:F34
    try:
        forma_pago_df = pd.read_excel(file_path, sheet_name="cta cobro", header=None, skiprows=32, nrows=2, usecols="B:F")
        forma_pago_text = " ".join([str(cell) for row in forma_pago_df.values for cell in row if pd.notna(cell)])
        df["FORMA DE PAGO"] = forma_pago_text.strip()
    except Exception as e:
        df["FORMA DE PAGO"] = ""
        print("[WARN] No se pudo leer FORMA DE PAGO:", e)

    # IVA TOTAL desde H34
    iva_total = str(ws["H34"].value).strip() if ws["H34"].value else ""
    df["IVA TOTAL"] = iva_total

    # CLIENTE desde B10:B15 (concatenado)
    cliente_partes = []
    for row in range(10, 16):
        val = ws[f"B{row}"].value
        if val:
            cliente_partes.append(str(val).strip())
    cliente = " ".join(cliente_partes)
    df["CLIENTE"] = cliente

    # ARCHIVO_ORIGEN con el nombre del archivo
    df["ARCHIVO_ORIGEN"] = os.path.basename(file_path)

    # Eliminar columnas completamente vacías
    df = df.dropna(axis=1, how='all')

    return df